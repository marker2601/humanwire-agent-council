"""Confined parsing tests for organization source uploads."""

from __future__ import annotations

import stat
import subprocess
import sys
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import pytest
from openpyxl import Workbook
from pypdf import PdfWriter
from pypdf._page import PageObject
from pypdf.generic import (
    ArrayObject,
    DecodedStreamObject,
    DictionaryObject,
    NameObject,
    NumberObject,
    TextStringObject,
)

from humanwire import organization_sources
from humanwire.organization_models import SourceRecord
from humanwire.organization_sources import (
    OrganizationSourceLimits,
    OrganizationSourceRejected,
    ParseOrganizationSourceRequest,
    parse_organization_source,
)

FIXTURES = Path(__file__).parents[1] / "fixtures" / "humanwire" / "organization"
ORG_ID = "org_01J00000000000000000000000"
EXPECTED_DIGEST = "9c18ef1f6480bb4ca6c69be44d086362853d8d1d31dde8892e8ad5e6c0fe2541"
MIME_TYPES = {
    ".csv": "text/csv",
    ".json": "application/json",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".pdf": "application/pdf",
}


def expected_source_records() -> tuple[SourceRecord, ...]:
    return (
        SourceRecord(
            record_id="rec_00000000000000000000000001",
            source_ordinal=1,
            source_identity="person:ada",
            fields=(("display_name", "Ada Lovelace"), ("title", "Chief Architect")),
        ),
        SourceRecord(
            record_id="rec_00000000000000000000000002",
            source_ordinal=2,
            source_identity="person:grace",
            fields=(("display_name", "Grace Hopper"), ("title", "Engineering Lead")),
        ),
    )


def source_request(
    fixture: str,
    *,
    content: bytes | None = None,
    filename: str | None = None,
    content_type: str | None = None,
    limits: OrganizationSourceLimits | None = None,
) -> ParseOrganizationSourceRequest:
    suffix = Path(fixture).suffix.lower()
    return ParseOrganizationSourceRequest(
        content=(FIXTURES / fixture).read_bytes() if content is None else content,
        filename=fixture if filename is None else filename,
        content_type=MIME_TYPES[suffix] if content_type is None else content_type,
        organization_id=ORG_ID,
        limits=limits or OrganizationSourceLimits(),
    )


def xlsx_bytes(rows: tuple[tuple[str, ...], ...]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def xlsx_with_formula(formula: str) -> ParseOrganizationSourceRequest:
    content = xlsx_bytes(
        (
            ("source_identity", "display_name", "title"),
            ("person:ada", "Ada Lovelace", formula),
        )
    )
    return source_request("sample.xlsx", content=content)


def xlsx_with_member(name: str, content: bytes) -> bytes:
    return rewrite_xlsx(additions=((name, content),))


def rewrite_xlsx(
    *,
    replacements: dict[str, bytes] | None = None,
    additions: tuple[tuple[str | ZipInfo, bytes], ...] = (),
) -> bytes:
    source = (FIXTURES / "sample.xlsx").read_bytes()
    output = BytesIO()
    with ZipFile(BytesIO(source)) as archive, ZipFile(output, "w", ZIP_DEFLATED) as target:
        for info in archive.infolist():
            value = (replacements or {}).get(info.filename, archive.read(info.filename))
            target.writestr(info, value)
        for name, content in additions:
            target.writestr(name, content)
    return output.getvalue()


def rewrite_sheet_xml(old: bytes, new: bytes) -> bytes:
    source = (FIXTURES / "sample.xlsx").read_bytes()
    with ZipFile(BytesIO(source)) as archive:
        sheet = archive.read("xl/worksheets/sheet1.xml")
    assert old in sheet
    return rewrite_xlsx(replacements={"xl/worksheets/sheet1.xml": sheet.replace(old, new, 1)})


def relationship_xml(*, target: str, relationship_id: str = "rId999") -> bytes:
    source = (FIXTURES / "sample.xlsx").read_bytes()
    with ZipFile(BytesIO(source)) as archive:
        relationships = archive.read("xl/_rels/workbook.xml.rels")
    addition = (
        f'<Relationship Id="{relationship_id}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        f'Target="{target}"/>'
    ).encode()
    return relationships.replace(b"</Relationships>", addition + b"</Relationships>")


def utf16_entity_content_types() -> bytes:
    source = (FIXTURES / "sample.xlsx").read_bytes()
    with ZipFile(BytesIO(source)) as archive:
        content_types = archive.read("[Content_Types].xml").decode()
    hostile = content_types.replace(
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
        '<?xml version="1.0" encoding="UTF-16"?>'
        '<!DOCTYPE Types [<!ENTITY extension "xml">]>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
        1,
    ).replace('Extension="xml"', 'Extension="&extension;"', 1)
    return hostile.encode("utf-16")


def multi_sheet_xlsx() -> bytes:
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(("source_identity", "display_name"))
    first.append(("person:one", "Person One"))
    second = workbook.create_sheet("Second")
    second.append(("source_identity", "display_name"))
    second.append(("person:two", "Person Two"))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def pdf_bytes(
    *,
    pages: int = 1,
    encrypted: bool = False,
    attachment: bool = False,
    javascript: bool = False,
) -> bytes:
    writer = PdfWriter()
    for _ in range(pages):
        writer.add_blank_page(width=72, height=72)
    if attachment:
        writer.add_attachment("private.txt", b"private")
    if javascript:
        writer.add_js("app.alert('private')")
    if encrypted:
        writer.encrypt("private-password")
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def pdf_with_decoded_content(size: int) -> bytes:
    writer = PdfWriter()
    page = writer.add_blank_page(width=72, height=72)
    stream = DecodedStreamObject()
    stream.set_data(b" " * size)
    page[NameObject("/Contents")] = writer._add_object(stream.flate_encode())
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def pdf_with_metadata(title: str) -> bytes:
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    writer.add_metadata({"/Title": title})
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def pdf_with_form() -> bytes:
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    form = DictionaryObject({NameObject("/Fields"): ArrayObject()})
    writer._root_object[NameObject("/AcroForm")] = writer._add_object(form)
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def pdf_with_cyclic_contents() -> bytes:
    writer = PdfWriter()
    page = writer.add_blank_page(width=72, height=72)
    contents = ArrayObject()
    contents_reference = writer._add_object(contents)
    contents.append(contents_reference)
    page[NameObject("/Contents")] = contents_reference
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def pdf_with_nested_form(decoded_size: int) -> bytes:
    writer = PdfWriter()
    page = writer.add_blank_page(width=72, height=72)

    form = DecodedStreamObject()
    form.set_data(b" " * decoded_size)
    form[NameObject("/Type")] = NameObject("/XObject")
    form[NameObject("/Subtype")] = NameObject("/Form")
    form[NameObject("/BBox")] = ArrayObject(
        [NumberObject(0), NumberObject(0), NumberObject(72), NumberObject(72)]
    )
    form[NameObject("/Resources")] = DictionaryObject()
    form_reference = writer._add_object(form.flate_encode())
    page[NameObject("/Resources")] = DictionaryObject(
        {NameObject("/XObject"): DictionaryObject({NameObject("/Fm1"): form_reference})}
    )

    content = DecodedStreamObject()
    content.set_data(b"q /Fm1 Do Q")
    page[NameObject("/Contents")] = writer._add_object(content)
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def pdf_with_many_safe_strings(count: int) -> bytes:
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    writer._root_object[NameObject("/SafeValues")] = ArrayObject(
        TextStringObject("safe") for _ in range(count)
    )
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def exception_graph_text(error: BaseException) -> str:
    pending = [error]
    seen: set[int] = set()
    rendered: list[str] = []
    while pending:
        current = pending.pop()
        if id(current) in seen:
            continue
        seen.add(id(current))
        rendered.extend((str(current), repr(current), repr(current.args)))
        if current.__cause__ is not None:
            pending.append(current.__cause__)
        if current.__context__ is not None:
            pending.append(current.__context__)
    return " ".join(rendered)


def private_graph_text(value: object, seen: set[int] | None = None) -> str:
    visited = set() if seen is None else seen
    if id(value) in visited:
        return ""
    visited.add(id(value))
    if isinstance(value, bytes):
        return repr(value)
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        return " ".join(
            private_graph_text(item, visited) for pair in value.items() for item in pair
        )
    if isinstance(value, (list, tuple, set, frozenset)):
        return " ".join(private_graph_text(item, visited) for item in value)
    return repr(value)


@pytest.mark.parametrize("fixture", ["sample.csv", "sample.json", "sample.xlsx"])
def test_structured_sources_produce_equal_canonical_rows(fixture: str) -> None:
    snapshot = parse_organization_source(source_request(fixture))

    assert snapshot.records == expected_source_records()
    assert snapshot.semantic_digest == EXPECTED_DIGEST
    assert snapshot.organization_id == ORG_ID


def test_pdf_produces_bounded_fragments_not_structured_people() -> None:
    snapshot = parse_organization_source(source_request("sample.pdf"))

    assert snapshot.source_kind == "pdf_fragment"
    assert snapshot.records
    assert all(record.fields[0][0] == "text_fragment" for record in snapshot.records)
    assert all(len(record.fields) == 1 for record in snapshot.records)
    assert all(len(record.fields[0][1]) <= 120 for record in snapshot.records)
    assert all(record.source_identity.startswith("pdf_page:") for record in snapshot.records)


@pytest.mark.parametrize(
    ("fixture", "content_type"),
    [
        ("sample.csv", "application/json"),
        ("sample.json", "text/csv"),
        ("sample.xlsx", "application/zip"),
        ("sample.pdf", "application/octet-stream"),
        ("sample.csv", "TEXT/CSV"),
        ("sample.csv", "text/csv; charset=utf-8"),
    ],
)
def test_extension_and_mime_must_match_exact_allowlist(
    fixture: str,
    content_type: str,
) -> None:
    with pytest.raises(OrganizationSourceRejected, match="^source_unsupported$"):
        parse_organization_source(source_request(fixture, content_type=content_type))


def test_extension_dispatch_is_lowercase_normalized() -> None:
    snapshot = parse_organization_source(source_request("sample.csv", filename="SAMPLE.CSV"))

    assert snapshot.records == expected_source_records()


def test_files_over_ten_mebibytes_are_rejected_before_parsing() -> None:
    content = b"x" * (10 * 1024 * 1024 + 1)

    with pytest.raises(OrganizationSourceRejected, match="^source_too_large$"):
        parse_organization_source(source_request("sample.csv", content=content))


@pytest.mark.parametrize(
    "content",
    [
        b"source_identity,display_name\nperson:one,\xff\n",
        b'source_identity,display_name\nperson:one,"unterminated\n',
        b"source_identity,display_name\nperson:one,Ada\x00Lovelace\n",
        b"source_identity,display_name\nperson:one,Ada\x01Lovelace\n",
        b"source_identity,display_name\nperson:one,=HYPERLINK(https://internal)\n",
        b"source_identity,display_name\nperson:one,+cmd|' /C private'!A0\n",
    ],
)
def test_csv_rejects_malformed_or_active_content(content: bytes) -> None:
    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.csv", content=content))


def test_csv_rejects_overlong_cells() -> None:
    content = f"source_identity,display_name\nperson:one,{'x' * 121}\n".encode()

    with pytest.raises(OrganizationSourceRejected, match="^source_too_large$"):
        parse_organization_source(source_request("sample.csv", content=content))


def test_csv_rejects_more_than_five_thousand_records() -> None:
    rows = ["source_identity,display_name"]
    rows.extend(f"person:{number},Person {number}" for number in range(5_001))

    with pytest.raises(OrganizationSourceRejected, match="^source_too_large$"):
        parse_organization_source(
            source_request("sample.csv", content=("\n".join(rows) + "\n").encode())
        )


def test_request_limits_may_reduce_but_not_raise_hard_boundaries() -> None:
    with pytest.raises(OrganizationSourceRejected, match="^source_too_large$"):
        parse_organization_source(
            source_request("sample.csv", limits=OrganizationSourceLimits(max_records=1))
        )

    with pytest.raises(ValueError, match="hard maximum"):
        OrganizationSourceLimits(max_records=5_001)


@pytest.mark.parametrize(
    "content",
    [
        b'{"source_identity":"person:one"}',
        b'[{"source_identity":"person:one","display_name":{"nested":"Ada"}}]',
        b'[{"source_identity":"person:one","display_name":42}]',
        b'[{"source_identity":"person:one","display_name":"Ada","display_name":"Grace"}]',
        b'[{"source_identity":"person:one","score":NaN}]',
        b'[{"source_identity":"person:one","score":Infinity}]',
        b'[{"source_identity":"person:one","score":-Infinity}]',
    ],
)
def test_json_rejects_duplicate_non_finite_or_invalid_structure(content: bytes) -> None:
    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.json", content=content))


def test_xlsx_formula_is_rejected() -> None:
    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(xlsx_with_formula('=HYPERLINK("https://internal")'))


def test_xlsx_formula_outside_declared_dimension_is_rejected() -> None:
    hostile = rewrite_sheet_xml(
        b"</sheetData>",
        b'<row r="100"><c r="D100"><f>HYPERLINK(&quot;https://internal&quot;)</f>'
        b"<v>1</v></c></row></sheetData>",
    )

    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.xlsx", content=hostile))


@pytest.mark.parametrize(
    "formula_xml",
    [
        b'<evil:f xmlns:evil="urn:evil">1+1</evil:f>',
        b'<f t="shared" si="0">1+1</f>',
        b'<f t="array" ref="A2:A2">1+1</f>',
        b'<f t="dataTable" ref="A2:A2">1+1</f>',
    ],
)
def test_xlsx_rejects_formula_local_name_in_every_namespace_and_form(
    formula_xml: bytes,
) -> None:
    hostile = rewrite_sheet_xml(
        b"<is><t>person:ada</t></is>",
        b"<is><t>person:ada</t></is>" + formula_xml,
    )

    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.xlsx", content=hostile))


def test_xlsx_rejects_foreign_worksheet_namespace_elements() -> None:
    hostile = rewrite_sheet_xml(
        b"<is><t>person:ada</t></is>",
        b'<is><t>person:ada</t></is><evil:payload xmlns:evil="urn:evil">safe</evil:payload>',
    )

    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.xlsx", content=hostile))


@pytest.mark.parametrize("dimension", [b"A1:XFD1048576", b"A1:IW3"])
def test_xlsx_unsafe_dimensions_are_rejected_before_openpyxl(
    dimension: bytes,
    monkeypatch,
) -> None:
    hostile = rewrite_sheet_xml(b"A1:C3", dimension)
    provider_called = False

    def fail_if_called(*_args, **_kwargs):
        nonlocal provider_called
        provider_called = True
        raise AssertionError("unsafe worksheet reached openpyxl")

    monkeypatch.setattr(organization_sources.openpyxl, "load_workbook", fail_if_called)

    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.xlsx", content=hostile))

    assert not provider_called


def test_xlsx_legitimate_multi_sheet_records_keep_global_ordinals() -> None:
    snapshot = parse_organization_source(source_request("sample.xlsx", content=multi_sheet_xlsx()))

    assert [record.source_ordinal for record in snapshot.records] == [1, 2]
    assert [record.source_identity for record in snapshot.records] == [
        "person:one",
        "person:two",
    ]


@pytest.mark.parametrize(
    ("member", "content"),
    [
        ("../private.txt", b"private"),
        ("xl/vbaProject.bin", b"macro"),
        (
            "custom/_rels/unsafe.rels",
            (
                b'<?xml version="1.0"?><Relationships '
                b'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships/hyperlink" '
                b'Target="https://private.invalid" '
                b'TargetMode="External"/></Relationships>'
            ),
        ),
        (
            "custom/_rels/unsupported.rels",
            (
                b'<?xml version="1.0"?><Relationships '
                b'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships/oleObject" '
                b'Target="../private.bin"/></Relationships>'
            ),
        ),
    ],
)
def test_xlsx_rejects_traversal_macros_and_unsafe_relationships(
    member: str,
    content: bytes,
) -> None:
    hostile = xlsx_with_member(member, content)

    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.xlsx", content=hostile))


def test_xlsx_rejects_zip_bombs() -> None:
    hostile = xlsx_with_member("xl/bomb.bin", b"0" * (1024 * 1024))

    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.xlsx", content=hostile))


def unsafe_xlsx_archives() -> tuple[bytes, ...]:
    symlink = ZipInfo("xl/symlink.xml")
    symlink.create_system = 3
    symlink.external_attr = (stat.S_IFLNK | 0o777) << 16
    malformed_relationships = (
        b'<NotRelationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'
    )
    entity_relationships = (
        b'<!DOCTYPE Relationships [<!ENTITY private "missing.xml">]>'
        b'<Relationships xmlns="http://schemas.openxmlformats.org/'
        b'package/2006/relationships"></Relationships>'
    )
    wrong_content_types = (
        b'<Types xmlns="https://private.invalid/content-types">'
        b'<Default Extension="xml" ContentType="application/xml"/></Types>'
    )
    return (
        rewrite_xlsx(additions=(("XL/WORKBOOK.XML", b"collision"),)),
        rewrite_xlsx(additions=((symlink, b"../workbook.xml"),)),
        rewrite_xlsx(additions=(("customXml/item1.xml", b"<private/>"),)),
        rewrite_xlsx(additions=(("xl/oleObject1.bin", b"private"),)),
        rewrite_xlsx(additions=(("custom/_rels/empty.rels", malformed_relationships),)),
        rewrite_xlsx(additions=(("custom/_rels/entity.rels", entity_relationships),)),
        rewrite_xlsx(
            replacements={
                "xl/_rels/workbook.xml.rels": relationship_xml(target="%2e%2e/%2e%2e/private.xml")
            }
        ),
        rewrite_xlsx(
            replacements={
                "xl/_rels/workbook.xml.rels": relationship_xml(target="worksheets/missing.xml")
            }
        ),
        rewrite_xlsx(replacements={"[Content_Types].xml": wrong_content_types}),
    )


@pytest.mark.parametrize("hostile", unsafe_xlsx_archives())
def test_xlsx_rejects_ambiguous_members_and_malformed_package_graph(hostile: bytes) -> None:
    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.xlsx", content=hostile))


def test_xlsx_rejects_utf16_entity_content_types() -> None:
    hostile = rewrite_xlsx(replacements={"[Content_Types].xml": utf16_entity_content_types()})

    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.xlsx", content=hostile))


@pytest.mark.parametrize("relationship_id", ["1invalid", "bad id", "bad:id"])
def test_xlsx_rejects_non_xml_relationship_ids(relationship_id: str) -> None:
    hostile = rewrite_xlsx(
        replacements={
            "xl/_rels/workbook.xml.rels": relationship_xml(
                target="worksheets/sheet1.xml",
                relationship_id=relationship_id,
            )
        }
    )

    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.xlsx", content=hostile))


@pytest.mark.parametrize(
    "content",
    [
        pdf_bytes(pages=65),
        pdf_bytes(encrypted=True),
        pdf_bytes(attachment=True),
        pdf_bytes(javascript=True),
    ],
)
def test_pdf_rejects_over_page_encrypted_embedded_and_active_files(content: bytes) -> None:
    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.pdf", content=content))


def test_pdf_rejects_cumulative_decoded_content_over_input_ceiling() -> None:
    hostile = pdf_with_decoded_content(10 * 1024 * 1024 + 1)
    assert len(hostile) < 10 * 1024 * 1024

    with pytest.raises(OrganizationSourceRejected, match="^source_too_large$"):
        parse_organization_source(source_request("sample.pdf", content=hostile))


def test_pdf_rejects_cyclic_contents_without_unbounded_traversal(tmp_path: Path) -> None:
    source = tmp_path / "cyclic.pdf"
    source.write_bytes(pdf_with_cyclic_contents())
    probe = """
import sys
from pathlib import Path
from humanwire.organization_sources import (
    OrganizationSourceRejected,
    ParseOrganizationSourceRequest,
    parse_organization_source,
)
content = Path(sys.argv[1]).read_bytes()
request = ParseOrganizationSourceRequest(
    content=content,
    filename="cyclic.pdf",
    content_type="application/pdf",
    organization_id="org_01J00000000000000000000000",
)
try:
    parse_organization_source(request)
except OrganizationSourceRejected as error:
    raise SystemExit(0 if str(error) == "source_unsafe" else 2)
raise SystemExit(3)
"""
    try:
        completed = subprocess.run(
            [sys.executable, "-c", probe, str(source)],
            check=False,
            timeout=2,
        )
    except subprocess.TimeoutExpired:
        pytest.fail("cyclic PDF contents did not terminate within the traversal bound")

    assert completed.returncode == 0


def test_pdf_counts_nested_form_streams_against_decoded_budget() -> None:
    hostile = pdf_with_nested_form(10 * 1024 * 1024 + 1)
    assert len(hostile) < 10 * 1024 * 1024

    with pytest.raises(OrganizationSourceRejected, match="^source_too_large$"):
        parse_organization_source(source_request("sample.pdf", content=hostile))


def test_pdf_provider_text_budget_aborts_during_chunk_extraction(monkeypatch) -> None:
    emitted = 0

    def emit_many_chunks(self, *args, visitor_text=None, **kwargs):
        nonlocal emitted
        for _ in range(100):
            emitted += 1
            if visitor_text is not None:
                visitor_text("x" * 121, None, None, None, 12)
        return "x" * 12_100

    monkeypatch.setattr(PageObject, "extract_text", emit_many_chunks)

    with pytest.raises(OrganizationSourceRejected, match="^source_too_large$"):
        parse_organization_source(
            source_request(
                "sample.pdf",
                limits=OrganizationSourceLimits(max_records=1),
            )
        )

    assert emitted == 1


@pytest.mark.parametrize(
    "private_value",
    [
        "password=PRIVATE-PDF-METADATA",
        r"C:\Users\private\organization.pdf",
        "powershell Invoke-Expression private-command",
    ],
)
def test_pdf_rejects_unsafe_reachable_metadata_values(private_value: str) -> None:
    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(
            source_request("sample.pdf", content=pdf_with_metadata(private_value))
        )


def test_pdf_rejects_unsupported_form_surface() -> None:
    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.pdf", content=pdf_with_form()))


def test_pdf_object_budget_counts_safe_scalar_values() -> None:
    hostile = pdf_with_many_safe_strings(20_001)

    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$"):
        parse_organization_source(source_request("sample.pdf", content=hostile))


@pytest.mark.parametrize(
    "private_value",
    [
        "api_key=sk-private-organization-secret",
        r"C:\Users\private\organization.xlsx",
        "/home/private/organization.csv",
        "powershell Invoke-Expression private-command",
        "rm -rf /private/organization",
    ],
)
def test_embedded_credentials_private_paths_and_commands_are_rejected(
    private_value: str,
) -> None:
    content = f"source_identity,display_name\nperson:one,{private_value}\n".encode()

    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$") as captured:
        parse_organization_source(source_request("sample.csv", content=content))

    assert private_value not in exception_graph_text(captured.value)


def test_fixed_errors_do_not_retain_filename_path_or_parsed_values() -> None:
    private_filename = r"C:\Users\private\organization.csv"
    private_value = "password=private-password-sentinel"
    request = source_request(
        "sample.csv",
        content=f"source_identity,display_name\nperson:one,{private_value}\n".encode(),
        filename=private_filename,
    )

    with pytest.raises(OrganizationSourceRejected) as captured:
        parse_organization_source(request)

    rendered = exception_graph_text(captured.value)
    assert str(captured.value) in {
        "source_invalid",
        "source_unsafe",
        "source_unsupported",
        "source_too_large",
    }
    assert private_filename not in rendered
    assert private_value not in rendered
    assert captured.value.__cause__ is None
    assert captured.value.__context__ is None


def test_fixed_errors_clear_parser_traceback_locals() -> None:
    sentinel = "PRIVATE-TRACEBACK-LOCAL-SENTINEL"
    request = source_request(
        "sample.csv",
        content=f"source_identity,display_name\nperson:one,password={sentinel}\n".encode(),
    )

    with pytest.raises(OrganizationSourceRejected, match="^source_unsafe$") as captured:
        parse_organization_source(request)

    traceback = captured.value.__traceback__
    rendered: list[str] = []
    while traceback is not None:
        if Path(traceback.tb_frame.f_code.co_filename).name == "organization_sources.py":
            rendered.append(private_graph_text(traceback.tb_frame.f_locals))
        traceback = traceback.tb_next
    assert sentinel not in " ".join(rendered)
    assert captured.value.__cause__ is None
    assert captured.value.__context__ is None


def test_provider_exceptions_are_sealed_without_retained_sentinel(monkeypatch) -> None:
    sentinel = "PRIVATE-PDF-PROVIDER-SENTINEL"

    def fail_reader(*_args, **_kwargs):
        raise RuntimeError(sentinel)

    monkeypatch.setattr(organization_sources.pypdf, "PdfReader", fail_reader)

    with pytest.raises(OrganizationSourceRejected, match="^source_invalid$") as captured:
        parse_organization_source(source_request("sample.pdf"))

    assert sentinel not in exception_graph_text(captured.value)
    assert captured.value.__cause__ is None
    assert captured.value.__context__ is None
